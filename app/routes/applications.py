from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file, current_app
from flask_login import login_required, current_user
from app import db
from app.models.application import Application
from app.models.resume import Resume
from app.models.activity import Activity
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

applications_bp = Blueprint('applications', __name__)

def log_activity(user_id, action, application_id=None, icon='circle'):
    activity = Activity(user_id=user_id, application_id=application_id, action=action, icon=icon)
    db.session.add(activity)

@applications_bp.route('/')
@login_required
def index():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    status_filter = request.args.get('status', '')
    sort_by = request.args.get('sort', 'created_at')
    order = request.args.get('order', 'desc')
    view = request.args.get('view', 'table')

    query = Application.query.filter_by(user_id=current_user.id)

    if search:
        query = query.filter(
            (Application.company_name.ilike(f'%{search}%')) |
            (Application.role_name.ilike(f'%{search}%')) |
            (Application.location.ilike(f'%{search}%'))
        )

    if status_filter:
        query = query.filter_by(status=status_filter)

    sort_col = getattr(Application, sort_by, Application.created_at)
    if order == 'desc':
        query = query.order_by(sort_col.desc())
    else:
        query = query.order_by(sort_col.asc())

    pagination = query.paginate(page=page, per_page=10, error_out=False)
    applications = pagination.items
    resumes = Resume.query.filter_by(user_id=current_user.id).all()

    return render_template('applications/index.html',
                           applications=applications,
                           pagination=pagination,
                           search=search,
                           status_filter=status_filter,
                           sort_by=sort_by,
                           order=order,
                           view=view,
                           resumes=resumes,
                           status_choices=Application.STATUS_CHOICES,
                           today=datetime.utcnow().date())

@applications_bp.route('/new', methods=['GET', 'POST'])
@login_required
def new():
    resumes = Resume.query.filter_by(user_id=current_user.id).all()
    if request.method == 'POST':
        app_date = request.form.get('application_date')
        deadline = request.form.get('deadline')
        resume_id = request.form.get('resume_id') or None

        application = Application(
            user_id=current_user.id,
            company_name=request.form.get('company_name', '').strip(),
            role_name=request.form.get('role_name', '').strip(),
            location=request.form.get('location', '').strip(),
            package=request.form.get('package', '').strip(),
            application_date=datetime.strptime(app_date, '%Y-%m-%d').date() if app_date else None,
            deadline=datetime.strptime(deadline, '%Y-%m-%d').date() if deadline else None,
            job_url=request.form.get('job_url', '').strip(),
            status=request.form.get('status', 'applied'),
            notes=request.form.get('notes', '').strip(),
            resume_id=int(resume_id) if resume_id else None
        )
        db.session.add(application)
        db.session.flush()

        log_activity(current_user.id,
                     f'Applied to {application.company_name} for {application.role_name}',
                     application.id, 'briefcase')
        db.session.commit()
        flash('Application added successfully!', 'success')
        return redirect(url_for('applications.index'))

    return render_template('applications/form.html',
                           application=None,
                           resumes=resumes,
                           status_choices=Application.STATUS_CHOICES)

@applications_bp.route('/<int:app_id>')
@login_required
def detail(app_id):
    application = Application.query.filter_by(id=app_id, user_id=current_user.id).first_or_404()
    activities = Activity.query.filter_by(application_id=app_id)\
        .order_by(Activity.created_at.desc()).all()
    resumes = Resume.query.filter_by(user_id=current_user.id).all()
    return render_template('applications/detail.html',
                           application=application,
                           activities=activities,
                           resumes=resumes,
                           status_choices=Application.STATUS_CHOICES)

@applications_bp.route('/<int:app_id>/edit', methods=['GET', 'POST'])
@login_required
def edit(app_id):
    application = Application.query.filter_by(id=app_id, user_id=current_user.id).first_or_404()
    resumes = Resume.query.filter_by(user_id=current_user.id).all()

    if request.method == 'POST':
        old_status = application.status
        app_date = request.form.get('application_date')
        deadline = request.form.get('deadline')
        resume_id = request.form.get('resume_id') or None

        application.company_name = request.form.get('company_name', '').strip()
        application.role_name = request.form.get('role_name', '').strip()
        application.location = request.form.get('location', '').strip()
        application.package = request.form.get('package', '').strip()
        application.application_date = datetime.strptime(app_date, '%Y-%m-%d').date() if app_date else None
        application.deadline = datetime.strptime(deadline, '%Y-%m-%d').date() if deadline else None
        application.job_url = request.form.get('job_url', '').strip()
        application.status = request.form.get('status', 'applied')
        application.notes = request.form.get('notes', '').strip()
        application.resume_id = int(resume_id) if resume_id else None
        application.updated_at = datetime.utcnow()

        if old_status != application.status:
            log_activity(current_user.id,
                         f'{application.company_name} status changed to {application.status_label}',
                         application.id, 'arrow-right')

        db.session.commit()
        flash('Application updated successfully!', 'success')
        return redirect(url_for('applications.detail', app_id=app_id))

    return render_template('applications/form.html',
                           application=application,
                           resumes=resumes,
                           status_choices=Application.STATUS_CHOICES)

@applications_bp.route('/<int:app_id>/delete', methods=['POST'])
@login_required
def delete(app_id):
    application = Application.query.filter_by(id=app_id, user_id=current_user.id).first_or_404()
    company = application.company_name
    db.session.delete(application)
    db.session.commit()
    flash(f'Application for {company} deleted.', 'info')
    return redirect(url_for('applications.index'))

@applications_bp.route('/export')
@login_required
def export():
    applications = Application.query.filter_by(user_id=current_user.id)\
        .order_by(Application.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Applications'

    headers = ['Company', 'Role', 'Location', 'Package', 'Status',
               'Application Date', 'Deadline', 'Job URL', 'Notes']

    header_fill = PatternFill(start_color='00FF88', end_color='00FF88', fill_type='solid')
    header_font = Font(bold=True, color='000000')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row, app in enumerate(applications, 2):
        ws.cell(row=row, column=1, value=app.company_name)
        ws.cell(row=row, column=2, value=app.role_name)
        ws.cell(row=row, column=3, value=app.location)
        ws.cell(row=row, column=4, value=app.package)
        ws.cell(row=row, column=5, value=app.status_label)
        ws.cell(row=row, column=6, value=str(app.application_date) if app.application_date else '')
        ws.cell(row=row, column=7, value=str(app.deadline) if app.deadline else '')
        ws.cell(row=row, column=8, value=app.job_url or '')
        ws.cell(row=row, column=9, value=app.notes or '')

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name='jobtrack_applications.xlsx')
